import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_row
def getColumnCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_column
def readData(file, sheetName, rowNo, coulmnNo) -> str:
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.cell(rowNo, coulmnNo).value
def writeData(file, sheetName, rowNo, coulmnNo,data):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    sheet.cell(rowNo, coulmnNo).value = data
    workBook.save(file)

def fillGreenColor(file, sheetName, rowNo, coulmnNo):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowNo,coulmnNo).fill = greenFill
    workBook.save(file)

def fillRedColor(file, sheetName, rowNo, coulmnNo):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNo,coulmnNo).fill = redFill
    workBook.save(file)
